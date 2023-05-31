from pathlib import Path
from typing import Optional, Any, Dict, List, TextIO
from threading import Thread


import PySimpleGUI as GUI
from openpyxl.worksheet.worksheet import Worksheet
from pydispatch import dispatcher

from core.gui import GraphicalUserInterface
from core.utils import ProcessTypes, match_extension, save_data, load_sheet, get_calculation_limits, convert_to_file_path
from core.sensors import Sensor
from core.exceptions import ValidationError, InvalidValueError
from core import signals
from core import settings


class Configurator(GraphicalUserInterface):
    def __init__(self):
        super().__init__()
        self.min_row: Optional[int] = None
        self.max_row: Optional[int] = None
        self.sheet: Optional[Worksheet] = None
        self.buffer: Path = Path(__file__).parent / 'buffer.txt'
        self.file = None
        self.is_processing: bool = True
        self.error_counter: int = 0
        self.sensors: Dict[str, List[Sensor]] = {}
        self.init_signals()

    def init_signals(self):
        super().init_signals()

    def handle_start_processing(self, sender: TextIO, **kwargs) -> None:
        self.file = sender
        self.is_processing = True
        super().handle_start_processing(sender, **kwargs)
        process_type = ProcessTypes(self.process_type_dropdown.get())
        self.file.write(process_type.start_string)

    def handle_sensor_processing(self, sender: Sensor, **kwargs) -> None:
        self.error_counter = 0
        self.sensors.setdefault(sender.__class__.__name__, []).append(sender)
        super().handle_sensor_processing(sender, self.error_counter, **kwargs)

    def handle_stop_processing(self, sender: Any, **kwargs) -> None:
        self.is_processing = False
        super().handle_stop_processing(sender, **kwargs)
        process_type = ProcessTypes(self.process_type_dropdown.get())

        for sensor_type, sensor_list in self.sensors.items():
            sensor_class = sensor_list[0].__class__
            cluster = sensor_class.clusterize(sensor_list, process_type)
            self.file.write(cluster)

        self.file.write(process_type.end_string)
        self.file.close()
        self.sensors.clear()
        self.file = None

    def handle_validation_failure(self, sender: Exception, **kwargs) -> None:
        self.error_counter += 1
        super().handle_validation_failure(sender, self.error_counter, **kwargs)
        if self.error_counter >= settings.MAX_FAILURES_PER_RUN:
            dispatcher.send(signals.STOP_PROCESSING,
                            reason=f'превышено максимальное количество ошибок ({settings.MAX_FAILURES_PER_RUN}).')

    def run(self):
        while True:
            event, values = self.window.read()

            # Загрузка данных
            if event == self.load_data_btn.key:
                file_path = GUI.popup_get_file(message='Загрузить данные',
                                               file_types=settings.SUPPORTED_LOAD_FILE_TYPES,
                                               no_window=True)
                if not file_path:
                    continue

                file_path = convert_to_file_path(file_path)
                self.sheet = load_sheet(file_path)
                if not self.sheet:
                    dispatcher.send(signals.DATA_LOADING_FAILED, file_path=file_path)
                    continue

                dispatcher.send(signals.DATA_LOADED, file_path=file_path)

            # Обработка данных
            elif event == self.process_data_btn.key:
                try:
                    min_row, max_row = get_calculation_limits(self.sheet, self.min_row, self.max_row)

                except InvalidValueError as exc:
                    dispatcher.send(signals.INVALID_INPUT_VALUE, exc)
                    continue

                Thread(target=self.process_data, args=(min_row, max_row)).start()

            # Сохранение данных
            elif event == self.save_data_btn.key:
                process_type = ProcessTypes(self.process_type_dropdown.get())
                file_path = GUI.popup_get_file(message='Сохранить данные',
                                               save_as=True,
                                               file_types=settings.SUPPORTED_SAVE_FILE_TYPES[process_type],
                                               no_window=True)
                if not file_path:
                    continue

                file_path = convert_to_file_path(file_path, extension=match_extension(process_type))
                save_data(self.buffer, file_path)
                dispatcher.send(signals.DATA_SAVED, file_path=file_path)

            # Остановка обработки данных
            elif event == self.stop_process_btn.key:
                dispatcher.send(signals.STOP_PROCESSING,
                                reason='Обработка прервана пользователем.')

            # Получение минимального и максимального номеров строк
            elif event in (self.min_row_input.key, self.max_row_input.key):
                self.min_row = values[self.min_row_input.key]
                self.max_row = values[self.max_row_input.key]

            # Побочное
            elif event == 'Об авторе':
                GUI.popup(event, settings.ABOUT_POPUP_TEXT)

            elif event == 'Инструкция':
                GUI.popup(event, settings.INSTRUCTION_POPUP_TEXT)

            elif event == GUI.WIN_CLOSED:
                break

        self.window.close()

    # Внутренний метод для обработки данных
    def process_data(self, min_row: int, max_row: int):
        with open(self.buffer, 'w', encoding='utf-8') as omx_file:
            dispatcher.send(signals.START_PROCESSING, omx_file, min_row=min_row, max_row=max_row)

            for row in range(min_row, max_row + 1):
                if not self.is_processing:
                    break

                skip_flag = self.sheet[f'{settings.SKIP_FLAG_COLUMN}{row}'].value
                if skip_flag not in (0, 1):
                    dispatcher.send(signals.STOP_PROCESSING,
                                    reason=f'Неверное значение флага пропуска на строке {row}.')
                    break

                if skip_flag == 0:
                    dispatcher.send(signals.SKIP_FLAG_DETECTED, row=row)
                    continue

                try:
                    sensor = Sensor.create(self.sheet, row)
                    dispatcher.send(signals.SENSOR_DETECTED, sensor, row=row)

                except ValidationError as exc:
                    dispatcher.send(signals.VALIDATION_FAILED, exc, row=row)

            if self.is_processing:
                dispatcher.send(signals.STOP_PROCESSING, reason='Обработка успешно завершена.')


if __name__ == '__main__':
    program = Configurator()
    program.run()
